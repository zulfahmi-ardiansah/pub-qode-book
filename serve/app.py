"""Qodebook — a read-only reader for the UNSPSC/KBLI SQLite database.

Pages:
  /                        pick a taxonomy
  /browse/{name}           the hierarchy: lazy tree + search
  /browse/{name}/{code}    one code: its position, definition, links, siblings
  /download/{what}.{ext}   a whole table as CSV or XLSX

Every connection is opened read-only, so nothing here can mutate the database
built by process/mapper.py.

Language: the UI runs in English or Indonesian, chosen by the `lang` cookie. Level
names are ALWAYS taken from LEVELS below, never from the database's `category`
column — that column holds English for UNSPSC ("Commodity") and Indonesian for
KBLI ("Kelompok"), which would otherwise mix two languages on one screen.
"""

import csv
import io
import os
import sqlite3
from functools import lru_cache, partial
from pathlib import Path
from tempfile import NamedTemporaryFile

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from starlette.background import BackgroundTask

SERVE_DIR = Path(__file__).resolve().parent
ROOT_DIR = SERVE_DIR.parent

load_dotenv(ROOT_DIR / '.env')

# Relative SQLITE_PATH is read from the project root, so the app runs the same
# whatever directory uvicorn was started in.
SQLITE_PATH = ROOT_DIR / os.getenv('SQLITE_PATH', 'data/database.sqlite')

LANGS = ('en', 'id')
DEFAULT_LANG = 'en'

# Level names per taxonomy per language. The Indonesian KBLI ladder is the
# official one; its English column uses the ISIC equivalents that KBLI is built
# on (Section / Division / Group / Class / Sub-class), not a literal translation.
LEVELS = {
    'unspsc': {
        'en': {0: 'Root', 1: 'Segment', 2: 'Family', 3: 'Class', 4: 'Commodity'},
        'id': {0: 'Akar', 1: 'Segmen', 2: 'Famili', 3: 'Kelas', 4: 'Komoditas'},
    },
    'kbli': {
        'en': {0: 'Section', 1: 'Division', 2: 'Group', 3: 'Class', 4: 'Sub-class'},
        'id': {0: 'Kategori', 1: 'Golongan Pokok', 2: 'Golongan', 3: 'Subgolongan', 4: 'Kelompok'},
    },
}

# Plural forms, used in running text ("10 classes below"). Indonesian does not
# mark plurals, so those forms are identical to the singular by design.
PLURALS = {
    'unspsc': {
        'en': {0: 'roots', 1: 'segments', 2: 'families', 3: 'classes', 4: 'commodities'},
        'id': {0: 'akar', 1: 'segmen', 2: 'famili', 3: 'kelas', 4: 'komoditas'},
    },
    'kbli': {
        'en': {0: 'sections', 1: 'divisions', 2: 'groups', 3: 'classes', 4: 'sub-classes'},
        'id': {0: 'kategori', 1: 'golongan pokok', 2: 'golongan', 3: 'subgolongan', 4: 'kelompok'},
    },
}

DATASETS = {
    'unspsc': {
        'table': 'master_unspsc',
        'label': 'UNSPSC',
        'version': '260801',
        'map_column': 'code_unspsc',
        # map_master only holds UNSPSC families (8-digit, xxxx0000).
        'map_level': 2,
        'specimen': '10151600',
        'specimen_level': 3,
    },
    'kbli': {
        'table': 'master_kbli',
        'label': 'KBLI',
        'version': '2025',
        'map_column': 'code_kbli',
        # map_master only holds KBLI kelompok (5-digit).
        'map_level': 4,
        'specimen': '01111',
        'specimen_level': 4,
    },
}

# The three things the download popup offers. `stem` names the file that lands on
# disk. The taxonomy exports lead with level and sort by it, so the file opens
# broad-to-specific — the hierarchy the app draws, flattened into rows. map_view
# is the mapping with both titles already joined on, so it reads without a lookup.
DOWNLOADS = {
    'unspsc': {
        'source': 'master_unspsc',
        'columns': 'level, code, parent_code, title, definition',
        'order': 'level, code',
        'stem': 'unspsc-260801',
        'sheet': 'UNSPSC',
    },
    'kbli': {
        'source': 'master_kbli',
        'columns': 'level, code, parent_code, title, definition',
        'order': 'level, code',
        'stem': 'kbli-2025',
        'sheet': 'KBLI',
    },
    'mapping': {
        'source': 'map_view',
        'columns': 'code_kbli, title_kbli, code_unspsc, title_unspsc',
        'order': 'code_kbli, code_unspsc',
        'stem': 'unspsc-kbli-mapping',
        'sheet': 'Mapping',
    },
}


def export_query(spec: dict) -> str:
    return f'SELECT {spec["columns"]} FROM {spec["source"]} ORDER BY {spec["order"]}'

FORMATS = {
    'csv': 'text/csv; charset=utf-8',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}

# The spreadsheet is read by people, so its header row says what a column means
# rather than what the database calls it. The CSV keeps the raw column names —
# that file is for machines.
COLUMN_LABELS = {
    'en': {
        'code': 'Code',
        'title': 'Title',
        'definition': 'Definition',
        'category': 'Level name',
        'parent_code': 'Parent code',
        'level': 'Level',
        'description': 'Description',
        'code_kbli': 'KBLI code',
        'title_kbli': 'KBLI activity',
        'code_unspsc': 'UNSPSC code',
        'title_unspsc': 'UNSPSC product or service',
    },
    'id': {
        'code': 'Kode',
        'title': 'Judul',
        'definition': 'Uraian',
        'category': 'Nama tingkat',
        'parent_code': 'Kode induk',
        'level': 'Tingkat',
        'description': 'Keterangan',
        'code_kbli': 'Kode KBLI',
        'title_kbli': 'Aktivitas KBLI',
        'code_unspsc': 'Kode UNSPSC',
        'title_unspsc': 'Barang atau jasa UNSPSC',
    },
}

COLUMN_WIDTHS = {
    'code': 14, 'title': 46, 'definition': 80, 'category': 16,
    'parent_code': 14, 'level': 8, 'description': 80,
    'code_kbli': 12, 'title_kbli': 46, 'code_unspsc': 14, 'title_unspsc': 46,
}

STRINGS = {
    'en': {
        'lang_name': 'English',
        'home_eyebrow': 'What this is',
        'home_title': 'Look Up The Code, See What It Links To.',
        'home_sub': 'Find the UNSPSC or KBLI code you need, whether you start from a product, a service, or a line of business. Look one up on either side, and this shows you the codes it maps to on the other.',
        'unspsc_tagline': 'What gets bought and sold',
        'unspsc_blurb': 'The United Nations Standard Products and Services Code (UNSPSC) is a global classification of goods and services. The UNDP started it with Dun & Bradstreet, and GS1 US maintains it today. Buyers and suppliers use it to say exactly what is being purchased, so spending stays comparable across catalogues, tenders, and borders. It holds around 158,000 entries across four levels: segment, family, class, and commodity.',
        'kbli_tagline': 'What a business does',
        'kbli_blurb': 'Klasifikasi Baku Lapangan Usaha Indonesia (KBLI) is the national classification of business activity. Badan Pusat Statistik (BPS) publishes it, built on the UN ISIC standard. Companies register their line of business under it, and BPS uses it to compile economic statistics, so every business in the country is counted the same way. It holds around 2,400 entries across five levels, from a two-digit main group down to a five-digit kelompok.',
        'codes': 'codes',
        'browse': 'Browse',
        'bridge_title': 'How the two sides are linked',
        'bridge_body': [
            'A KBLI code tells you what a company does. A UNSPSC code tells you what it sells. They are separate systems built for separate purposes, so nothing joins them on its own.', 
            'This app carries {pairs} links, each connecting a five digit KBLI kelompok to a UNSPSC family whose goods or services that activity plausibly produces or sells.', 
            'One activity usually delivers many products, and one product can come from many activities, so the two sides meet many to many rather than one to one.'
        ],
        'warning': 'Warning',
        'bridge_caveat': 'The links were generated with a large language model (Gemini 3.5 Flash), matching UNSPSC at the family level to KBLI at the kelompok level. Some of them maybe are wrong, so check any link you intend to rely on.',
        'links_to': 'links to',
        'see_it': 'See it',
        'search_ph': 'Search {total} codes, titles and definitions',
        'hierarchy': 'The hierarchy',
        'matches': 'Matches',
        'none': 'none',
        'first': 'first {n}',
        'no_match': 'No code, title or definition contains “{q}”.',
        'searching': 'Searching…',
        'loading': 'Loading…',
        'broad': 'Broad',
        'specific': 'Specific',
        'by_level': 'Codes at each level',
        'definition': 'Definition',
        'no_definition': 'No definition recorded for this code. The title above is all the source gives.',
        'read_full': 'Read the full definition',
        'show_less': 'Show less',
        'n_more': '{n} more',
        'same_parent': 'Same parent',
        'under': 'Under {code}',
        'top_of_hierarchy': 'Top of the hierarchy',
        'no_siblings': 'This code sits at the top of the hierarchy, so it has no siblings.',
        'position': 'Position',
        'you_are_here': 'You are here',
        'leaf': 'a leaf — nothing below it',
        'n_below': '{n} {level} below',
        'n_links': '{n} link{s} to {label}',
        'no_link': 'No {label} link recorded for this code.',
        'facts_parent': 'Parent',
        'facts_below': 'Below',
        'facts_links': 'Links',
        'facts_level': 'Level',
        'none_dash': '—',
        'unspsc_mapping_heading': 'Activities that produce or sell this',
        'kbli_mapping_heading': 'Products and services this activity delivers',
        'inherited': 'Inherited from family {code}',
        'rolled_up': 'Rolled up from the sub-classes under {code}',
        'filter': 'Filter {n} {noun}',
        'nothing_matches': 'Nothing here matches “{q}”.',
        'prev': 'Prev',
        'next': 'Next',
        'of': 'of',
        'show': 'Show',
        'download': 'Download',
        'download_title': 'Take the data with you',
        'download_sub': 'Three tables, exactly as the database holds them. Pick a format.',
        'download_close': 'Close',
        'download_rows': '{n} rows',
        'download_working': 'Preparing…',
        'download_failed': 'That download failed. Try again.',
        'download_unspsc': 'Every UNSPSC code with its title, definition, parent and level.',
        'download_kbli': 'Every KBLI code with its title, definition, parent and level.',
        'download_mapping': 'One row per link: the KBLI code and the UNSPSC code it connects to, each with its title.',
        'mapping': 'Mapping',
        'links': 'links',
    },
    'id': {
        'lang_name': 'Bahasa Indonesia',
        'home_eyebrow': 'Tentang alat ini',
        'home_title': 'Cari Kodenya, Lihat Pemetaannya.',
        'home_sub': 'Temukan kode UNSPSC atau KBLI yang Anda butuhkan, entah mulai dari sebuah produk, jasa, atau bidang usaha. Cari dari sisi mana pun, dan halaman ini menunjukkan kode pasangannya di sisi seberang.',
        'unspsc_tagline': 'Apa yang diperjualbelikan',
        'unspsc_blurb': 'United Nations Standard Products and Services Code (UNSPSC) adalah klasifikasi global untuk barang dan jasa. UNDP memulainya bersama Dun & Bradstreet, dan kini GS1 US yang memeliharanya. Pembeli dan pemasok memakainya untuk menyebut persis apa yang dibeli, sehingga belanja tetap bisa dibandingkan antar katalog, tender, dan negara. Isinya sekitar 158.000 entri dalam empat tingkat: segmen, famili, kelas, dan komoditas.',
        'kbli_tagline': 'Apa yang dikerjakan usaha',
        'kbli_blurb': 'Klasifikasi Baku Lapangan Usaha Indonesia (KBLI) adalah klasifikasi nasional untuk aktivitas usaha. Badan Pusat Statistik (BPS) yang menerbitkannya, dengan struktur mengikuti standar ISIC milik PBB. Perusahaan mendaftarkan bidang usahanya memakai kode ini, dan BPS memakainya untuk menyusun statistik ekonomi, sehingga seluruh usaha di Indonesia terhitung dengan cara yang sama. Isinya sekitar 2.400 entri dalam lima tingkat, dari golongan pokok dua digit sampai kelompok lima digit.',
        'codes': 'kode',
        'browse': 'Telusuri',
        'bridge_title': 'Bagaimana keduanya dipetakan',
        'bridge_body': [
            'Kode KBLI memberi tahu apa yang dikerjakan sebuah usaha. Kode UNSPSC memberi tahu apa yang dijualnya. Keduanya sistem terpisah yang dibangun untuk tujuan berbeda, jadi tidak ada yang menyambungkannya dengan sendirinya.',
            'Aplikasi ini memuat {pairs} pemetaan, masing-masing menghubungkan satu kelompok KBLI lima digit dengan famili UNSPSC yang barang atau jasanya masuk akal dihasilkan atau dijual oleh aktivitas tersebut.',
            'Satu aktivitas biasanya menghasilkan banyak produk, dan satu produk bisa datang dari banyak aktivitas, jadi keduanya bertemu banyak ke banyak, bukan satu ke satu.'
        ],
        'warning': 'Perhatian',
        'bridge_caveat': 'Pemetaan ini dibuat dengan bantuan model bahasa besar (Gemini 3.5 Flash), yang mencocokkan UNSPSC di tingkat famili dengan KBLI di tingkat kelompok, jadi bukan pemetaan resmi. Sebagian di antaranya mungkin keliru, jadi periksa dulu pemetaan yang hendak Anda pakai.',
        'links_to': 'dipetakan ke',
        'see_it': 'Lihat',
        'search_ph': 'Cari {total} kode, judul, dan uraian',
        'hierarchy': 'Hierarki',
        'matches': 'Hasil',
        'none': 'nihil',
        'first': '{n} pertama',
        'no_match': 'Tidak ada kode, judul, atau uraian yang memuat “{q}”.',
        'searching': 'Mencari…',
        'loading': 'Memuat…',
        'broad': 'Luas',
        'specific': 'Spesifik',
        'by_level': 'Kode di tiap tingkat',
        'definition': 'Uraian',
        'no_definition': 'Tidak ada uraian untuk kode ini. Hanya judul di atas yang tersedia di sumbernya.',
        'read_full': 'Baca uraian lengkap',
        'show_less': 'Tutup',
        'n_more': '{n} lainnya',
        'same_parent': 'Satu induk',
        'under': 'Di bawah {code}',
        'top_of_hierarchy': 'Puncak hierarki',
        'no_siblings': 'Kode ini berada di puncak hierarki, jadi tidak punya saudara.',
        'position': 'Posisi',
        'you_are_here': 'Anda di sini',
        'leaf': 'tidak ada turunan',
        'n_below': '{n} {level} di bawahnya',
        'n_links': '{n} pemetaan ke {label}',
        'no_link': 'Tidak ada pemetaan {label} untuk kode ini.',
        'facts_parent': 'Induk',
        'facts_below': 'Turunan',
        'facts_links': 'Pemetaan',
        'facts_level': 'Tingkat',
        'none_dash': '—',
        'unspsc_mapping_heading': 'Aktivitas yang menghasilkan atau menjual ini',
        'kbli_mapping_heading': 'Barang dan jasa yang dihasilkan aktivitas ini',
        'inherited': 'Diwarisi dari famili {code}',
        'rolled_up': 'Dirangkum dari kelompok di bawah {code}',
        'filter': 'Saring {n} {noun}',
        'nothing_matches': 'Tidak ada yang cocok dengan “{q}”.',
        'prev': 'Sebelumnya',
        'next': 'Berikutnya',
        'of': 'dari',
        'show': 'Menampilkan',
        'download': 'Unduh',
        'download_title': 'Bawa datanya',
        'download_sub': 'Tiga tabel, persis seperti yang tersimpan di basis data. Pilih formatnya.',
        'download_close': 'Tutup',
        'download_rows': '{n} baris',
        'download_working': 'Menyiapkan…',
        'download_failed': 'Unduhan gagal. Coba lagi.',
        'download_unspsc': 'Seluruh kode UNSPSC beserta judul, uraian, induk, dan tingkatnya.',
        'download_kbli': 'Seluruh kode KBLI beserta judul, uraian, induk, dan tingkatnya.',
        'download_mapping': 'Satu baris per pemetaan: kode KBLI dan kode UNSPSC yang terhubung dengannya, masing-masing dengan judulnya.',
        'mapping': 'Pemetaan',
        'links': 'pemetaan',
    },
}

SEARCH_LIMIT = 100
CSV_CHUNK = 64 * 1024

HEADER_FONT = Font(bold=True, color='14181F')
HEADER_FILL = PatternFill('solid', fgColor='E9EBF0')
HEADER_ALIGN = Alignment(vertical='center')

app = FastAPI(title='Qodebook')
app.mount('/static', StaticFiles(directory=SERVE_DIR / 'static'), name='static')
templates = Jinja2Templates(directory=SERVE_DIR / 'templates')


# --------------------------------------------------------------------------- language

def lang_of(request: Request) -> str:
    value = request.cookies.get('lang', DEFAULT_LANG)
    return value if value in LANGS else DEFAULT_LANG


def level_name(name: str, level: int, lang: str) -> str:
    return LEVELS[name][lang].get(level, '')


def level_plural(name: str, level: int, lang: str) -> str:
    return PLURALS[name][lang].get(level, '')


def page_context(request: Request, **extra) -> dict:
    """Shared context: the language, its strings, and level-name helpers."""
    lang = lang_of(request)
    strings = STRINGS[lang]
    return {
        'lang': lang,
        'langs': LANGS,
        't': strings,
        'level_name': partial(level_name, lang=lang),
        'level_plural': partial(level_plural, lang=lang),
        # The download popup lives in the masthead, so every page needs its numbers.
        'downloads': DOWNLOADS,
        'download_counts': download_counts(),
        # Level names for the client, so JS renders the tree in the same language.
        'js_levels': {n: [LEVELS[n][lang][i] for i in range(5)] for n in LEVELS},
        'js_strings': {
            k: strings[k] for k in
            ('loading', 'searching', 'none', 'first', 'no_match', 'nothing_matches',
             'filter', 'prev', 'next', 'of', 'read_full', 'show_less', 'n_more', 'show',
             'download_working', 'download_failed')
        },
        **extra,
    }


@app.get('/lang/{code}')
def set_lang(code: str, request: Request):
    """Switch language and return to the page you were on."""
    if code not in LANGS:
        raise HTTPException(404, f'Unknown language {code!r}')
    back = request.headers.get('referer') or '/'
    response = RedirectResponse(back, status_code=303)
    response.set_cookie('lang', code, max_age=60 * 60 * 24 * 365, samesite='lax')
    return response


# --------------------------------------------------------------------------- db

def connect() -> sqlite3.Connection:
    """Open the database read-only, so no request can mutate what mapper.py built."""
    if not SQLITE_PATH.exists():
        raise HTTPException(500, f'Database not found at {SQLITE_PATH}')
    conn = sqlite3.connect(f'file:{SQLITE_PATH.as_posix()}?mode=ro', uri=True)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA query_only = 1')
    return conn


@lru_cache(maxsize=1)
def download_counts() -> dict[str, int]:
    """Row counts for the download popup. The database never changes at runtime,
    so count once rather than on every page render."""
    with connect() as conn:
        return {
            key: conn.execute(f'SELECT COUNT(*) FROM {spec["source"]}').fetchone()[0]
            for key, spec in DOWNLOADS.items()
        }


def dataset(name: str) -> dict:
    if name not in DATASETS:
        raise HTTPException(404, f'Unknown dataset {name!r}')
    return DATASETS[name]


def other(name: str) -> str:
    return 'kbli' if name == 'unspsc' else 'unspsc'


def rows_to_dicts(cursor) -> list[dict]:
    return [dict(r) for r in cursor.fetchall()]


def node_rows(conn, table: str, where: str, params: tuple) -> list[dict]:
    """Nodes plus a has_children flag, so the tree knows which rows can expand."""
    return rows_to_dicts(conn.execute(
        f'SELECT code, title, level, parent_code, '
        f'       EXISTS (SELECT 1 FROM {table} c WHERE c.parent_code = t.code) AS has_children '
        f'FROM {table} t WHERE {where} ORDER BY code', params
    ))


# --------------------------------------------------------------------------- code shape

def code_slots(name: str, code: str, level: int) -> list[dict]:
    """Break a code into the level slots its digits encode.

    This is the reading key for both taxonomies. A UNSPSC code is four digit pairs
    — segment, family, class, commodity — and a pair of `00` means "this level is
    not specified". KBLI starts at two digits and gains one per level. Slots past
    the node's own level come back unfilled, which is what makes a code legible as
    a position rather than a number.
    """
    if level == 0:
        return [{'digits': code, 'level': 0, 'filled': True, 'current': True}]

    widths = [2, 2, 2, 2] if name == 'unspsc' else [2, 1, 1, 1]
    slots, cursor = [], 0
    for index, width in enumerate(widths):
        depth = index + 1                     # slot 0 is level 1
        digits = code[cursor:cursor + width]
        cursor += width
        slots.append({
            'digits': digits or '·' * width,
            'level': depth,
            'filled': depth <= level,
            'current': depth == level,
        })
    return slots


def code_split(name: str, code: str, level: int) -> tuple[str, str]:
    """Split a code into meaningful digits and trailing placeholder zeros.

    Lets a list render `1015` bright and `0000` dim, so a column of UNSPSC codes
    shows depth without reading a single digit.
    """
    if name != 'unspsc' or level < 1:
        return code, ''
    return code[:level * 2], code[level * 2:]


templates.env.globals.update(code_slots=code_slots, code_split=code_split, datasets=DATASETS)


# --------------------------------------------------------------------------- pages

@app.get('/', response_class=HTMLResponse)
def home(request: Request):
    lang = lang_of(request)
    with connect() as conn:
        cards = []
        for name, ds in DATASETS.items():
            cards.append({
                'name': name,
                **ds,
                'tagline': STRINGS[lang][f'{name}_tagline'],
                'blurb': STRINGS[lang][f'{name}_blurb'],
                'total': conn.execute(f'SELECT COUNT(*) FROM {ds["table"]}').fetchone()[0],
            })
        pairs = conn.execute('SELECT COUNT(*) FROM map_master').fetchone()[0]
    return templates.TemplateResponse(
        request, 'home.html', page_context(request, cards=cards, pairs=pairs)
    )


@app.get('/browse/{name}', response_class=HTMLResponse)
def browse_page(request: Request, name: str):
    ds = dataset(name)
    lang = lang_of(request)
    with connect() as conn:
        roots = node_rows(conn, ds['table'], 'level = 0', ())
        counts = dict(conn.execute(
            f'SELECT level, COUNT(*) FROM {ds["table"]} GROUP BY level'
        ).fetchall())
        total = sum(counts.values())
    return templates.TemplateResponse(request, 'browse.html', page_context(
        request, name=name, ds=ds, roots=roots, total=total, counts=counts,
        tagline=STRINGS[lang][f'{name}_tagline'], blurb=STRINGS[lang][f'{name}_blurb'],
    ))


@app.get('/browse/{name}/{code}', response_class=HTMLResponse)
def detail_page(request: Request, name: str, code: str):
    ds = dataset(name)
    lang = lang_of(request)
    peer_name = other(name)
    peer = DATASETS[peer_name]

    with connect() as conn:
        row = conn.execute(
            f'SELECT code, title, definition, parent_code, level, description '
            f'FROM {ds["table"]} WHERE code = ?', (code,)
        ).fetchone()
        if row is None:
            raise HTTPException(404, f'{ds["label"]} code {code!r} not found')
        node = dict(row)

        children = node_rows(conn, ds['table'], 'parent_code = ?', (code,))
        siblings = (
            node_rows(conn, ds['table'], 'parent_code = ?', (node['parent_code'],))
            if node['parent_code'] else []
        )
        mappings = _mappings(conn, name, node, lang)

        ancestors, parent = [], node['parent_code']
        while parent and len(ancestors) < 8:
            found = conn.execute(
                f'SELECT code, title, level FROM {ds["table"]} WHERE code = ?', (parent,)
            ).fetchone()
            if found is None:
                break
            ancestors.append(dict(found))
            parent = conn.execute(
                f'SELECT parent_code FROM {ds["table"]} WHERE code = ?', (parent,)
            ).fetchone()[0]
        ancestors.reverse()

    return templates.TemplateResponse(request, 'detail.html', page_context(
        request, name=name, ds=ds, peer_name=peer_name, peer=peer,
        node=node, ancestors=ancestors, children=children, siblings=siblings,
        mappings=mappings,
        definition=node['definition'] or node['description'],
        mapping_heading=STRINGS[lang][f'{name}_mapping_heading'],
        slots=code_slots(name, node['code'], node['level']),
    ))


# --------------------------------------------------------------------------- download

def csv_rows(query: str):
    """Stream a whole table as CSV, a chunk at a time.

    UNSPSC is 158k rows, so the file is built as it is sent rather than assembled
    in memory first. The leading BOM is what makes Excel read the Indonesian
    titles as UTF-8 instead of mangling them.
    """
    conn = connect()
    try:
        cursor = conn.execute(query)
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator='\n')
        buffer.write('﻿')
        writer.writerow([d[0] for d in cursor.description])
        for row in cursor:
            writer.writerow(list(row))
            if buffer.tell() >= CSV_CHUNK:
                yield buffer.getvalue()
                buffer.seek(0)
                buffer.truncate(0)
        yield buffer.getvalue()
    finally:
        conn.close()


def xlsx_path(query: str, sheet: str, lang: str) -> str:
    """Write a whole table to a temp .xlsx and return its path.

    The header row is bold, filled and frozen, and reads in the language the page
    was in. openpyxl's write-only workbook keeps one row in memory at a time; the
    caller deletes the file once the response has been sent.
    """
    labels = COLUMN_LABELS[lang]
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet)
    conn = connect()
    try:
        cursor = conn.execute(query)
        columns = [d[0] for d in cursor.description]

        for index, column in enumerate(columns, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 18)
        # Scrolling 158k rows is useless if the header scrolls away with them.
        worksheet.freeze_panes = 'A2'

        header = []
        for column in columns:
            cell = WriteOnlyCell(worksheet, value=labels.get(column, column))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            header.append(cell)
        worksheet.append(header)

        for row in cursor:
            worksheet.append(list(row))
    finally:
        conn.close()

    handle = NamedTemporaryFile(suffix='.xlsx', delete=False)
    handle.close()
    workbook.save(handle.name)
    return handle.name


@app.get('/download/{what}.{ext}')
def download(request: Request, what: str, ext: str):
    """One table, one file. `what` is a key of DOWNLOADS, `ext` is csv or xlsx."""
    if what not in DOWNLOADS:
        raise HTTPException(404, f'Nothing to download called {what!r}')
    if ext not in FORMATS:
        raise HTTPException(404, f'Unknown format {ext!r}')

    spec = DOWNLOADS[what]
    filename = f'{spec["stem"]}.{ext}'
    disposition = {'Content-Disposition': f'attachment; filename="{filename}"'}

    if ext == 'csv':
        return StreamingResponse(
            csv_rows(export_query(spec)), media_type=FORMATS[ext], headers=disposition
        )

    path = xlsx_path(export_query(spec), spec['sheet'], lang_of(request))
    return FileResponse(
        path, media_type=FORMATS[ext], filename=filename,
        background=BackgroundTask(os.unlink, path),
    )


# --------------------------------------------------------------------------- api

@app.get('/api/tree/{name}')
def tree(name: str, parent: str):
    """Children of one node — the tree fetches these lazily as branches open."""
    ds = dataset(name)
    with connect() as conn:
        return {'rows': node_rows(conn, ds['table'], 'parent_code = ?', (parent,))}


@app.get('/api/search/{name}')
def search(name: str, q: str = ''):
    """Flat search across code, title and definition."""
    ds = dataset(name)
    q = q.strip()
    if not q:
        return {'rows': [], 'total': 0}

    params = {'like': f'%{q}%'}
    where = '(code LIKE :like OR title LIKE :like OR definition LIKE :like)'
    with connect() as conn:
        total = conn.execute(f'SELECT COUNT(*) FROM {ds["table"]} WHERE {where}', params).fetchone()[0]
        rows = rows_to_dicts(conn.execute(
            f'SELECT code, title, level FROM {ds["table"]} WHERE {where} '
            f'ORDER BY level, code LIMIT {SEARCH_LIMIT}', params
        ))
    return {'rows': rows, 'total': total, 'limit': SEARCH_LIMIT}


def _mappings(conn, name: str, node: dict, lang: str) -> dict:
    """Cross-taxonomy matches for a node.

    map_master only links UNSPSC families to KBLI kelompok. For a node above or
    below that level we fall back to the nearest linked relative and say so, so
    the page is never blank for a code that is in fact covered.
    """
    ds, peer = DATASETS[name], DATASETS[other(name)]
    select = (
        f'SELECT DISTINCT p.code, p.title, p.level '
        f'FROM map_master m JOIN {peer["table"]} p ON p.code = m.{peer["map_column"]} '
    )
    direct = rows_to_dicts(conn.execute(
        select + f'WHERE m.{ds["map_column"]} = ? ORDER BY p.code', (node['code'],)
    ))
    if direct:
        return {'items': direct, 'note': None}

    code, level = node['code'], node['level']
    if name == 'unspsc' and level > ds['map_level']:
        family = code[:4] + '0000'
        inherited = rows_to_dicts(conn.execute(
            select + f'WHERE m.{ds["map_column"]} = ? ORDER BY p.code', (family,)
        ))
        note = STRINGS[lang]['inherited'].format(code=family) if inherited else None
        return {'items': inherited, 'note': note}

    if name == 'kbli' and 1 <= level < ds['map_level']:
        rolled = rows_to_dicts(conn.execute(
            select + f'WHERE m.{ds["map_column"]} LIKE ? ORDER BY p.code', (f'{code}%',)
        ))
        note = STRINGS[lang]['rolled_up'].format(code=code) if rolled else None
        return {'items': rolled, 'note': note}

    return {'items': [], 'note': None}
